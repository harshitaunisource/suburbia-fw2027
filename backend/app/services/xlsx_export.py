"""
Shared Excel-export helper. Both the Market Analytics and Buyer
Opportunities pages need a "Download Excel" button -- this keeps the
actual openpyxl workbook-building code in one place instead of
duplicated per router.

Each function returns raw .xlsx bytes (never writes to disk itself) so
the calling router can stream it straight back as a FileResponse-style
download without managing temp files.
"""
from __future__ import annotations

import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font

HEADER_FONT = Font(bold=True)


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 60)


def _write_table(ws, headers: list[str], rows: list[list]):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for row in rows:
        ws.append(row)
    _autosize(ws)


def build_analytics_workbook(report: dict, category: Optional[str], sources: Optional[list[str]]) -> bytes:
    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary.append(["Market Analytics Export"])
    summary["A1"].font = Font(bold=True, size=14)
    summary.append(["Category", category or "All"])
    summary.append(["Brands", ", ".join(sources) if sources else "All"])
    _autosize(summary)

    ws_price = wb.create_sheet("Price Distribution")
    _write_table(
        ws_price,
        ["Currency", "Min (MRP)", "Avg (MRP)", "Max (MRP)", "Product Count"],
        [
            [currency, stats["min"], stats["avg"], stats["max"], stats["count"]]
            for currency, stats in report["price_distribution"].items()
        ],
    )

    ws_counts = wb.create_sheet("Product Counts")
    _write_table(
        ws_counts,
        ["Source", "Product Count"],
        [[source, count] for source, count in report["product_counts"]["by_source"].items()],
    )

    for sheet_name, key in [
        ("Top Colors", "colors"),
        ("Silhouette-Fit", "silhouettes"),
        ("Patterns", "patterns"),
        ("Necklines", "necklines"),
    ]:
        ws = wb.create_sheet(sheet_name)
        _write_table(ws, ["Value", "Count"], [[row["value"], row["count"]] for row in report[key]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_opportunities_workbook(
    opportunities: list, gap_table: list[dict], category: str, our_source: str, competitor_sources: Optional[list[str]]
) -> bytes:
    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary.append(["Buyer Opportunities Export"])
    summary["A1"].font = Font(bold=True, size=14)
    summary.append(["Category", category])
    summary.append(["Buyer", our_source])
    summary.append(["Competitors", ", ".join(competitor_sources) if competitor_sources else "All others"])
    _autosize(summary)

    ws_opp = wb.create_sheet("Opportunities")
    _write_table(
        ws_opp,
        [
            "Concept", "Opportunity Score", "Trend", "Competitor", "Gap",
            "Price", "Commercial", "Status", "Reason",
        ],
        [
            [
                o.concept_name, o.opportunity_score, o.trend_score, o.competitor_score,
                o.suburbia_gap_score, o.price_score, o.commercial_score, o.status, o.reason,
            ]
            for o in opportunities
        ],
    )

    ws_gap = wb.create_sheet("Gap Table")
    _write_table(
        ws_gap,
        ["Concept", "Market %", "Buyer %", "Gap", "Gap Level"],
        [
            [row.get("value"), row.get("market_pct"), row.get("suburbia_pct"), row.get("gap"), row.get("gap_label")]
            for row in gap_table
        ],
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()